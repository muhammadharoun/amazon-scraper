import sys
from colorama import Fore, init, Back, Style
import openpyxl
import re

init(convert=True)
print("\n")
path = input("Enter xls file path, ex- C:\\employee.xlsx : ")
input_col_name = input("Enter colname, ex- Endpoint : ")
try:
    print(Fore.RESET)
    #path = "C:\\employee.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    # from the active attribute 
    sheet_obj = wb_obj.active

    # get max column count
    max_column=sheet_obj.max_column
    max_row=sheet_obj.max_row
    for j in range(2, 5):
        salary_cell=sheet_obj.cell(row=j,column=3)
        print(salary_cell.value)
        if float(salary_cell.value) < 1500:
            salary_cell.value =  salary_cell.value+500

    wb_obj.save(path.strip())
except Exception as e:
    print(e)
    print (Fore.RED + "Error : The file does not found")
print(Fore.GREEN + "###################### Successfully! Excel file has been read/write. ##############################")
