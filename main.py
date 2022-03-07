# importing libraries
import os
import sys
import re
import openpyxl
import requests
import urllib.parse
import json
from time import sleep
from csv import writer
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from colorama import Fore, init, Back, Style
# from selenium.webdriver.common.s import Keys

#  lxml
# configurations


def add_titles(File, titles):
    writer_object = writer(File)
    writer_object.writerow(titles)


def get_amazon_products(URL, File, min_price, max_price, min_rate, min_review, max_review, amazon_link):  # ,
    # opening our output file in append mode

    HEADERS = ({'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64)  AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36',
               'Accept-Language': 'en-US, en;q=0.5'})

    # Making the HTTP Request
    webpage = requests.get(URL, headers=HEADERS)
    # Creating the Soup Object containing all data
    soup = BeautifulSoup(webpage.content, "lxml")
    try:
        cards = soup.find_all("div", attrs={"class": 's-card-container'})
        product_data = []
        for ele in cards:
            #  get title
            title = ele.find('h2').find('a').find('span').text
            # get rate
            try:
                rate = float(
                    ele.find('div', attrs={"class": 'a-row a-size-small'}).text.split(' ')[0])
            except:
                rate = 0

            # get price
            try:
                price = float(
                    ele.find('span', attrs={"class": 'a-offscreen'}).text.replace('ريال', ''))
            except:
                price = 0

            # get link
            try:
                link = amazon_link+ele.find('h2').find('a').get('href')
            except:
                link = ''

            # get review
            try:
                review = int(ele.find('a', attrs={
                             "class": 'a-link-normal s-underline-text s-underline-link-text s-link-style'}).find('span').text)
            except:
                review = 0

            product_data.append(title)
            product_data.append(price)
            product_data.append(review)
            product_data.append(rate)
            product_data.append(link)

            if (min_price < price < max_price) and (min_review < review < max_review) and min_rate < rate:
                writer_object = writer(File)
                writer_object.writerow(product_data)
            product_data = []
    except AttributeError:
        title_string = "NA"


def getConfigData(option):
    config = json.load(open('config.json', 'r', encoding='utf-8-sig'))

    if option == 'options':
        return config['options']['min_price'], config['options']['max_price'], config['options']['min_review'], config['options']['max_review'], config['options']['min_rate'], config['options']['keywords'], config['options']['amazon_link']
    elif option == 'external_options':
        return config["external_options"]['output_name'],config["external_options"]['update_file'], config["external_options"]['sleep'], config['external_options']['start_pages'],config['external_options']['end_pages']

# get product stock amazon use selenium
def getStock(URL,amazon_link):
    driver = webdriver.Chrome(f"{os.getcwd()}/chromedriver.exe")
    driver.get(URL)
    sleep(3) 
    driver.find_element(by=By.XPATH, value="//*[@id='add-to-cart-button']").click()
    driver.get(f'{amazon_link}/gp/cart/view.html')
    sleep(3)
    driver.find_element(by=By.XPATH, value="//*[@id='a-autoid-0-announce']").click()
    sleep(3)
    driver.find_element(by=By.XPATH, value="//*[@id='quantity_10']").click()
    sleep(3)
    qyt = driver.find_element(by=By.XPATH, value="//*[@class='a-input-text a-width-small a-spacing-mini sc-quantity-textfield sc-update-quantity-input sc-hidden']")
    qyt.send_keys('999')
    sleep(3)
    driver.find_element(by=By.XPATH, value="//*[@id='a-autoid-1-announce']").click()
    sleep(3)
    try:
        stock = driver.find_element(by=By.XPATH, value="//*[@class='a-input-text a-width-small a-spacing-mini sc-quantity-textfield sc-update-quantity-input']").get_attribute("value")
    except:
        try:
            stock = driver.find_element(by=By.XPATH, value="//*[@class='a-dropdown-prompt']").text
        except:
            stock = 'error'
            
    return stock


def updata_stock(update_file,update_day,amazon_link):
    init(convert=True)
    path = fr'{os.getcwd()}\{update_file}.xlsx'
    day_number = int(update_day) + 2
    try:
        print(Fore.RESET)
        wb_obj = openpyxl.load_workbook(path.strip())
        # from the active attribute 
        sheet_obj = wb_obj.active

        # get max column count
        max_column=sheet_obj.max_column
        max_row=sheet_obj.max_row
        for j in range(2, max_row+1):
            link=sheet_obj.cell(row=j,column=2)
            try:
                stock = getStock(link.value,amazon_link) # get stock  . . 
            except:
                stock = 'error'
            day=sheet_obj.cell(row=j,column=day_number)
            day.value = stock
        wb_obj.save(path.strip())
    except Exception as e:
        print(e)
        print (Fore.RED + "Error : The file does not found")
        
    print(" Successfully! Excel file has been update. ")



def run_code():
    min_price, max_price, min_review, max_review, min_rate, keywords, amazon_link = getConfigData(
        'options')
    output_name,update_file, sleep_time, start_pages,end_pages = getConfigData('external_options')

    print('''
    - select || 1 || if to catch product 
    - select || 2 || if to update data 
    ''')
    input_val = input('    # => ')
    if input_val == '1':
        File = open(f'{os.getcwd()}/{output_name}.csv', 'a', newline='', encoding='utf-8-sig')
        add_titles(File, ['title', 'price', 'review', 'rate', 'link'])
        keyword = urllib.parse.quote(keywords)
        for page in range(start_pages,end_pages):
            get_amazon_products(f'{amazon_link}/s?k={keyword}&page={page}', File,
                                min_price, max_price, min_rate, min_review, max_review, amazon_link)

            sleep(float(sleep_time))
        File.close()
    elif input_val == '2':
        update_day = input('update day > ')
        updata_stock(update_file,update_day,amazon_link)
    else:
        print('plz enter valid number')

    


if __name__ == '__main__':
    run_code()


# select section 1
# 1 - get all products
# 2 - save products in execl sheet

#  select section 2
# 1- read all products from execl sheet
# 2- updata data (add to card senario)


# cart ===>> https://www.amazon.sa/gp/cart/view.html
